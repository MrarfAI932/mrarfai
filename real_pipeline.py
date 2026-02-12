#!/usr/bin/env python3
"""
MRARFAI V9.0 — Sprocomm 真实数据分析管线
==========================================
读取 Sprocomm 01401.HK 实际出货数据
输出: sprocomm_real_results.json → 供 Dashboard 使用

数据源:
  ① 25年出货数据by金额报表_最终.xlsx  (收入, 万元)
  ② 手机_平板出货by数量2025_-_最终汇总.xlsx (出货量, 台)
"""

import openpyxl
import json
import os
from collections import defaultdict
from datetime import datetime

REV_FILE = "/mnt/user-data/uploads/25年出货数据by金额报表_最终.xlsx"
QTY_FILE = "/mnt/user-data/uploads/手机_平板出货by数量2025_-_最终汇总.xlsx"
OUTPUT = os.path.join(os.path.dirname(__file__), "sprocomm_real_results.json")


def read_revenue_data():
    """读取金额数据 (万元 RMB)"""
    wb = openpyxl.load_workbook(REV_FILE, data_only=True)

    # === 2025数据 sheet: 客户月度收入 ===
    ws = wb["2025数据"]
    customers = []
    for row in range(5, 43):  # Row 5-42 are real customers, 43+ are summaries
        sales_person = ws.cell(row, 1).value
        customer = ws.cell(row, 2).value
        total = ws.cell(row, 3).value

        if not customer or customer == "汇总":
            continue
        if not isinstance(total, (int, float)):
            continue
        if total == 0:
            continue  # Skip zero-revenue entries

        monthly = []
        for col in range(4, 16):  # 1月-12月
            v = ws.cell(row, col).value
            monthly.append(round(float(v), 2) if v and isinstance(v, (int, float)) else 0)

        customers.append({
            "sales_person": str(sales_person or "").strip(),
            "customer": str(customer).strip(),
            "total_rev": round(float(total), 2),
            "monthly_rev": monthly,
            "h1_rev": round(sum(monthly[:6]), 2),
            "h2_rev": round(sum(monthly[6:]), 2),
        })

    # === Sheet2: 业务类别汇总 ===
    ws2 = wb["Sheet2"]
    categories = []
    for row in range(2, ws2.max_row + 1):
        cat = ws2.cell(row, 1).value
        rev_2025 = ws2.cell(row, 2).value
        share_2025 = ws2.cell(row, 3).value
        rev_2024 = ws2.cell(row, 4).value
        share_2024 = ws2.cell(row, 5).value
        growth_amt = ws2.cell(row, 6).value
        growth_pct = ws2.cell(row, 7).value

        if cat and rev_2025 and cat != "汇总" and isinstance(rev_2025, (int, float)):
            categories.append({
                "category": str(cat).strip(),
                "rev_2025": round(float(rev_2025), 1),
                "share_2025": round(float(share_2025) * 100, 1) if share_2025 else 0,
                "rev_2024": round(float(rev_2024), 1) if rev_2024 else 0,
                "share_2024": round(float(share_2024) * 100, 1) if share_2024 else 0,
                "growth_amt": round(float(growth_amt), 1) if growth_amt else 0,
                "growth_pct": round(float(growth_pct) * 100, 1) if growth_pct else 0,
            })

    # === Sheet2 汇总行 ===
    total_row = None
    for row in range(2, ws2.max_row + 1):
        if ws2.cell(row, 1).value == "汇总":
            rev25 = ws2.cell(row, 2).value
            rev24 = ws2.cell(row, 4).value
            ga = ws2.cell(row, 6).value
            gp = ws2.cell(row, 7).value
            total_row = {
                "rev_2025": round(float(rev25), 1) if isinstance(rev25, (int, float)) else 0,
                "rev_2024": round(float(rev24), 1) if isinstance(rev24, (int, float)) else 0,
                "growth_amt": round(float(ga), 1) if isinstance(ga, (int, float)) else 0,
                "growth_pct": round(float(gp) * 100, 1) if isinstance(gp, (int, float)) else 0,
            }
            break

    # === Sheet3: 区域分布 ===
    ws3 = wb["Sheet3"]
    regions = []
    for row in range(2, ws3.max_row + 1):
        region = ws3.cell(row, 1).value
        amount = ws3.cell(row, 2).value
        if region and amount and isinstance(amount, (int, float)):
            regions.append({
                "region": str(region).strip(),
                "rev": round(float(amount), 1),
            })

    # === Sheet1: Q1/Q2 计划 vs 实际 ===
    ws1 = wb["Sheet1"]
    q1p = ws1.cell(2, 2).value
    q1a = ws1.cell(3, 2).value
    quarterly = {
        "q1_plan": round(float(q1p), 1) if isinstance(q1p, (int, float)) else 0,
        "q1_2024": round(float(q1a), 1) if isinstance(q1a, (int, float)) else 0,
    }

    return {
        "customers": customers,
        "categories": categories,
        "total": total_row,
        "regions": regions,
        "quarterly": quarterly,
    }


def read_quantity_data():
    """读取出货量数据 (台)"""
    wb = openpyxl.load_workbook(QTY_FILE, data_only=True)
    ws = wb["数量汇总"]

    customers = []
    totals = None
    product_mix = {}

    for row in range(4, ws.max_row + 1):
        customer = ws.cell(row, 2).value
        if not customer:
            continue

        customer = str(customer).strip()

        # Read plan/actual pairs for each month
        monthly = []
        for m in range(6):  # 6 months
            plan_col = 3 + m * 2
            actual_col = 4 + m * 2
            plan = ws.cell(row, plan_col).value
            actual = ws.cell(row, actual_col).value
            monthly.append({
                "month": m + 1,
                "plan": int(plan) if plan and isinstance(plan, (int, float)) else 0,
                "actual": int(actual) if actual and isinstance(actual, (int, float)) else 0,
            })

        if customer == "TOTAL" or customer == "Total":
            totals = monthly
        elif customer in ("FP", "PAD", "SP"):
            product_mix[customer] = monthly
        elif customer not in ("汇总", "合计", "Type", "功能机与智能机分布："):
            total_plan = sum(m["plan"] for m in monthly)
            total_actual = sum(m["actual"] for m in monthly)

            if total_plan > 0 or total_actual > 0:
                completion = round(total_actual / total_plan * 100, 1) if total_plan > 0 else 0
                customers.append({
                    "customer": customer,
                    "monthly": monthly,
                    "h1_plan": total_plan,
                    "h1_actual": total_actual,
                    "completion_pct": completion,
                    "gap": total_actual - total_plan,
                })

    return {
        "customers": customers,
        "totals": totals,
        "product_mix": product_mix,
    }


def detect_anomalies(rev_data, qty_data):
    """多维异常检测"""
    anomalies = []

    # 1. 收入月度环比异常 (MoM)
    for c in rev_data["customers"]:
        monthly = c["monthly_rev"]
        for i in range(1, 6):  # month 2-6
            prev, curr = monthly[i-1], monthly[i]
            if prev > 100:  # 排除低量噪声 (>100万)
                change = (curr / prev - 1) * 100 if prev else 0
                if abs(change) > 40:
                    anomalies.append({
                        "type": "revenue_mom",
                        "severity": "critical" if abs(change) > 60 else "warning",
                        "customer": c["customer"],
                        "month": i + 1,
                        "prev_val": round(prev, 0),
                        "curr_val": round(curr, 0),
                        "change_pct": round(change, 1),
                        "desc": f"{c['customer']} {i+1}月收入{'暴增' if change > 0 else '骤降'} {change:+.1f}% ({prev:.0f}→{curr:.0f}万元)",
                    })

    # 2. 出货量计划 vs 实际偏差
    for c in qty_data["customers"]:
        for m in c["monthly"]:
            if m["plan"] > 5000:
                deviation = (m["actual"] / m["plan"] - 1) * 100 if m["plan"] else 0
                if deviation < -40:
                    anomalies.append({
                        "type": "plan_miss",
                        "severity": "critical" if deviation < -60 else "warning",
                        "customer": c["customer"],
                        "month": m["month"],
                        "plan": m["plan"],
                        "actual": m["actual"],
                        "change_pct": round(deviation, 1),
                        "desc": f"{c['customer']} {m['month']}月出货远低于计划 {deviation:+.1f}% (计划{m['plan']:,} 实际{m['actual']:,})",
                    })
                elif deviation > 50:
                    anomalies.append({
                        "type": "plan_exceed",
                        "severity": "warning",
                        "customer": c["customer"],
                        "month": m["month"],
                        "plan": m["plan"],
                        "actual": m["actual"],
                        "change_pct": round(deviation, 1),
                        "desc": f"{c['customer']} {m['month']}月出货大幅超计划 {deviation:+.1f}% (计划{m['plan']:,} 实际{m['actual']:,})",
                    })

    # 3. H1 整体完成率异常
    for c in qty_data["customers"]:
        if c["h1_plan"] > 10000 and c["completion_pct"] < 60:
            anomalies.append({
                "type": "h1_underperform",
                "severity": "critical",
                "customer": c["customer"],
                "month": 0,
                "change_pct": round(c["completion_pct"] - 100, 1),
                "desc": f"{c['customer']} H1完成率仅 {c['completion_pct']:.1f}% (计划{c['h1_plan']:,} 实际{c['h1_actual']:,})",
            })

    anomalies.sort(key=lambda a: (0 if a["severity"] == "critical" else 1, -abs(a["change_pct"])))
    return anomalies[:25]


def generate_executive_summary(rev_data, qty_data, anomalies):
    """管理层摘要"""
    total = rev_data["total"]
    critical = [a for a in anomalies if a["severity"] == "critical"]
    warnings = [a for a in anomalies if a["severity"] == "warning"]

    # Top 5 客户
    top5 = sorted(rev_data["customers"], key=lambda c: c["total_rev"], reverse=True)[:5]

    # 总出货量
    if qty_data["totals"]:
        h1_plan = sum(m["plan"] for m in qty_data["totals"])
        h1_actual = sum(m["actual"] for m in qty_data["totals"])
        completion = round(h1_actual / h1_plan * 100, 1) if h1_plan else 0
    else:
        h1_plan = h1_actual = completion = 0

    return {
        "generated_at": datetime.now().isoformat(),
        "headline": f"2025年收入 {total['rev_2025']/10000:.1f}亿元, 同比+{total['growth_pct']:.1f}%, H1出货完成率 {completion}%",

        "key_metrics": {
            "rev_2025_total": total["rev_2025"],
            "rev_2024_total": total["rev_2024"],
            "yoy_growth_pct": total["growth_pct"],
            "yoy_growth_amt": total["growth_amt"],
            "h1_shipment_plan": h1_plan,
            "h1_shipment_actual": h1_actual,
            "h1_completion_pct": completion,
        },

        "key_findings": [
            f"2025年总收入 {total['rev_2025']:,.0f}万元 ({total['rev_2025']/10000:.2f}亿), YoY +{total['growth_pct']:.1f}%",
            f"2024年总收入 {total['rev_2024']:,.0f}万元, 净增长 {total['growth_amt']:,.0f}万元",
            f"H1出货计划 {h1_plan:,} 台, 实际完成 {h1_actual:,} 台, 完成率 {completion}%",
            f"发现 {len(critical)} 个严重异常, {len(warnings)} 个预警",
            f"Top客户: {', '.join(c['customer'] for c in top5)}",
        ],

        "top_customers": [
            {
                "name": c["customer"],
                "rev": c["total_rev"],
                "share": round(c["total_rev"] / total["rev_2025"] * 100, 1),
                "sales": c["sales_person"],
            }
            for c in top5
        ],

        "critical_alerts": [
            {"alert": a["desc"], "severity": a["severity"]}
            for a in critical[:5]
        ],

        "recommendations": generate_recommendations(rev_data, qty_data, anomalies),
    }


def generate_recommendations(rev_data, qty_data, anomalies):
    """基于数据生成建议"""
    recs = []

    # 检查 HMD 下降
    hmd = next((c for c in rev_data["customers"] if c["customer"] == "HMD"), None)
    hmd_cat = next((c for c in rev_data["categories"] if c["category"] == "HMD"), None)
    if hmd_cat and hmd_cat["growth_pct"] < 0:
        recs.append(f"⚠️ HMD 收入同比 {hmd_cat['growth_pct']:+.1f}%, 下降 {abs(hmd_cat['growth_amt']):,.0f}万元 — 建议与 HMD 团队确认订单前景及竞品动态")

    # 检查高增长品类
    for cat in rev_data["categories"]:
        if cat["growth_pct"] > 50:
            recs.append(f"📈 {cat['category']} 同比增长 +{cat['growth_pct']:.0f}% — 建议确认产能/供应链能否支撑持续增长")

    # 出货计划偏差
    plan_misses = [c for c in qty_data["customers"] if c["completion_pct"] < 70 and c["h1_plan"] > 50000]
    if plan_misses:
        names = ", ".join(c["customer"] for c in plan_misses[:3])
        recs.append(f"📉 {names} H1出货远低于计划 — 建议重新评估下半年目标并制定追赶方案")

    # 区域集中度
    if rev_data["regions"]:
        top_region = rev_data["regions"][0]
        total_region = sum(r["rev"] for r in rev_data["regions"])
        if total_region > 0:
            concentration = top_region["rev"] / total_region * 100
            if concentration > 50:
                recs.append(f"🌍 区域集中度过高: {top_region['region']}占比 {concentration:.0f}% — 建议加大其他区域开拓力度")

    recs.append("📊 建议启动H2旺季备货计划, 结合历史季节性数据优化排产节奏")

    return recs


def build_dashboard_data(rev_data, qty_data, anomalies, summary):
    """构建 Dashboard 所需数据"""

    # 月度收入趋势
    monthly_total = []
    month_names = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    for m in range(12):
        total = sum(c["monthly_rev"][m] for c in rev_data["customers"])
        monthly_total.append({
            "month": month_names[m],
            "revenue": round(total, 0),
        })

    # 客户排名
    customer_ranking = sorted(rev_data["customers"], key=lambda c: c["total_rev"], reverse=True)
    top_customers = []
    total_rev = rev_data["total"]["rev_2025"]
    for c in customer_ranking[:10]:
        top_customers.append({
            "name": c["customer"],
            "rev": round(c["total_rev"], 0),
            "share": round(c["total_rev"] / total_rev * 100, 1) if total_rev else 0,
            "sales": c["sales_person"],
            "h1": round(c["h1_rev"], 0),
            "h2": round(c["h2_rev"], 0),
        })

    # 出货量计划 vs 实际
    qty_comparison = []
    for c in sorted(qty_data["customers"], key=lambda x: x["h1_actual"], reverse=True)[:10]:
        qty_comparison.append({
            "name": c["customer"],
            "plan": c["h1_plan"],
            "actual": c["h1_actual"],
            "completion": c["completion_pct"],
            "gap": c["gap"],
        })

    # 月度出货 plan vs actual
    monthly_shipments = []
    if qty_data["totals"]:
        for m in qty_data["totals"]:
            monthly_shipments.append({
                "month": month_names[m["month"] - 1],
                "plan": m["plan"],
                "actual": m["actual"],
                "completion": round(m["actual"] / m["plan"] * 100, 1) if m["plan"] else 0,
            })

    # 产品组合 (FP/SP/PAD)
    product_mix = []
    for ptype, label in [("FP", "功能机"), ("SP", "智能机"), ("PAD", "平板")]:
        if ptype in qty_data["product_mix"]:
            total_actual = sum(m["actual"] for m in qty_data["product_mix"][ptype])
            product_mix.append({"type": label, "code": ptype, "qty": total_actual})

    return {
        "monthly_revenue": monthly_total,
        "top_customers": top_customers,
        "categories": rev_data["categories"],
        "regions": rev_data["regions"],
        "anomalies": anomalies[:12],
        "qty_comparison": qty_comparison,
        "monthly_shipments": monthly_shipments,
        "product_mix": product_mix,
        "summary": summary,
        "total": rev_data["total"],
    }


def main():
    print("=" * 60)
    print("  MRARFAI V9.0 — Sprocomm 真实数据分析")
    print("=" * 60)

    print("\n📂 读取金额数据...")
    rev_data = read_revenue_data()
    print(f"   客户: {len(rev_data['customers'])} 个")
    print(f"   类别: {len(rev_data['categories'])} 个")
    print(f"   区域: {len(rev_data['regions'])} 个")

    print("\n📂 读取出货量数据...")
    qty_data = read_quantity_data()
    print(f"   客户: {len(qty_data['customers'])} 个")

    print("\n🔍 异常检测...")
    anomalies = detect_anomalies(rev_data, qty_data)
    critical = [a for a in anomalies if a["severity"] == "critical"]
    print(f"   发现: {len(anomalies)} 个异常 ({len(critical)} critical)")

    print("\n📊 生成管理层摘要...")
    summary = generate_executive_summary(rev_data, qty_data, anomalies)

    print("\n📈 构建 Dashboard 数据...")
    dashboard = build_dashboard_data(rev_data, qty_data, anomalies, summary)

    # Save
    with open(OUTPUT, "w", encoding="utf-8") as fout:
        json.dump(dashboard, fout, ensure_ascii=False, indent=2, default=str)

    print(f"\n✅ 分析完成!")
    print(f"   输出: {OUTPUT}")
    print(f"   大小: {os.path.getsize(OUTPUT) / 1024:.1f} KB")

    # Print summary
    print(f"\n{'='*60}")
    print(f"  📊 {summary['headline']}")
    print(f"{'='*60}")
    for finding in summary["key_findings"]:
        print(f"  • {finding}")
    if summary["critical_alerts"]:
        print(f"\n  ⚠️ 严重异常:")
        for a in summary["critical_alerts"][:5]:
            print(f"    🔴 {a['alert']}")
    print(f"\n  💡 建议:")
    for r in summary["recommendations"][:4]:
        print(f"    {r}")


if __name__ == "__main__":
    main()
