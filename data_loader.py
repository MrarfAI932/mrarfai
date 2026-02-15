"""
MRARFAI V9.0 — 数据加载模块
从 Excel 文件或缓存 JSON 读取 Sprocomm 出货数据
"""

import json
import os
import hashlib
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

CACHE_DIR = Path(__file__).parent / ".cache"
CACHE_DIR.mkdir(exist_ok=True)


def _file_hash(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        h.update(f.read())
    return h.hexdigest()[:12]


def load_from_json(path: str) -> dict:
    """从 JSON 结果文件加载"""
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def parse_revenue_excel(path: str) -> dict:
    """解析金额报表 Excel"""
    wb = openpyxl.load_workbook(path, data_only=True)

    # 2025数据 sheet
    ws = wb["2025数据"]
    customers = []
    for row in range(5, ws.max_row + 1):
        sales = ws.cell(row, 1).value
        customer = ws.cell(row, 2).value
        total = ws.cell(row, 3).value
        if not customer or not total or customer == "汇总":
        if not isinstance(total, (int, float)):
            continue
            continue
        monthly = []
        for col in range(4, 16):
            v = ws.cell(row, col).value
            monthly.append(round(float(v), 2) if v and isinstance(v, (int, float)) else 0)
        customers.append({
            "sales_person": str(sales or "").strip(),
            "customer": str(customer).strip(),
            "total_rev": round(float(total), 2),
            "monthly_rev": monthly,
            "h1_rev": round(sum(monthly[:6]), 2),
            "h2_rev": round(sum(monthly[6:]), 2),
        })

    # Sheet2: 业务类别
    ws2 = wb["Sheet2"]
    categories = []
    total_row = None
    for row in range(2, ws2.max_row + 1):
        cat = ws2.cell(row, 1).value
        rev25 = ws2.cell(row, 2).value
        share25 = ws2.cell(row, 3).value
        rev24 = ws2.cell(row, 4).value
        share24 = ws2.cell(row, 5).value
        growth_amt = ws2.cell(row, 6).value
        growth_pct = ws2.cell(row, 7).value

        if cat == "汇总":
            total_row = {
                "rev_2025": round(float(rev25 or 0), 1),
                "rev_2024": round(float(rev24 or 0), 1),
                "growth_amt": round(float(growth_amt or 0), 1),
                "growth_pct": round(float(growth_pct or 0) * 100, 1),
            }
        elif cat and rev25:
            categories.append({
                "category": str(cat).strip(),
                "rev_2025": round(float(rev25), 1),
                "share_2025": round(float(share25 or 0) * 100, 1),
                "rev_2024": round(float(rev24 or 0), 1),
                "share_2024": round(float(share24 or 0) * 100, 1),
                "growth_amt": round(float(growth_amt or 0), 1),
                "growth_pct": round(float(growth_pct or 0) * 100, 1),
            })

    # Sheet3: 区域
    ws3 = wb["Sheet3"]
    regions = []
    for row in range(2, ws3.max_row + 1):
        region = ws3.cell(row, 1).value
        amount = ws3.cell(row, 2).value
        if region and amount:
            regions.append({"region": str(region).strip(), "rev": round(float(amount), 1)})

    return {
        "customers": customers,
        "categories": categories,
        "total": total_row,
        "regions": regions,
    }


def parse_quantity_excel(path: str) -> dict:
    """解析出货量 Excel"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["数量汇总"]

    customers = []
    totals = None
    product_mix = {}
    month_names = ["1月", "2月", "3月", "4月", "5月", "6月"]

    for row in range(4, ws.max_row + 1):
        customer = ws.cell(row, 2).value
        if not customer:
            continue
        customer = str(customer).strip()

        monthly = []
        for m in range(6):
            plan_col = 3 + m * 2
            actual_col = 4 + m * 2
            plan = ws.cell(row, plan_col).value
            actual = ws.cell(row, actual_col).value
            monthly.append({
                "month": m + 1,
                "plan": int(plan) if plan and isinstance(plan, (int, float)) else 0,
                "actual": int(actual) if actual and isinstance(actual, (int, float)) else 0,
            })

        if customer in ("TOTAL", "Total"):
            totals = monthly
        elif customer in ("FP", "PAD", "SP"):
            product_mix[customer] = monthly
        elif customer not in ("汇总", "合计", "Type", "功能机与智能机分布："):
            total_plan = sum(m["plan"] for m in monthly)
            total_actual = sum(m["actual"] for m in monthly)
            if total_plan > 0 or total_actual > 0:
                completion = round(total_actual / total_plan * 100, 1) if total_plan > 0 else 0
                customers.append({
                    "customer": customer,
                    "monthly": monthly,
                    "h1_plan": total_plan,
                    "h1_actual": total_actual,
                    "completion_pct": completion,
                    "gap": total_actual - total_plan,
                })

    return {"customers": customers, "totals": totals, "product_mix": product_mix}


def detect_anomalies(rev_data: dict, qty_data: dict) -> list:
    """多维异常检测"""
    anomalies = []

    # 收入环比异常
    for c in rev_data["customers"]:
        monthly = c["monthly_rev"]
        for i in range(1, 6):
            prev, curr = monthly[i-1], monthly[i]
            if prev > 100:
                change = (curr / prev - 1) * 100 if prev else 0
                if abs(change) > 40:
                    anomalies.append({
                        "type": "收入环比",
                        "severity": "🔴 严重" if abs(change) > 60 else "🟡 预警",
                        "customer": c["customer"],
                        "month": f"{i+1}月",
                        "change_pct": round(change, 1),
                        "detail": f"{prev:.0f} → {curr:.0f} 万元",
                        "direction": "up" if change > 0 else "down",
                    })

    # 出货计划偏差
    for c in qty_data["customers"]:
        for m in c["monthly"]:
            if m["plan"] > 5000:
                dev = (m["actual"] / m["plan"] - 1) * 100 if m["plan"] else 0
                if dev < -40:
                    anomalies.append({
                        "type": "出货偏差",
                        "severity": "🔴 严重" if dev < -60 else "🟡 预警",
                        "customer": c["customer"],
                        "month": f"{m['month']}月",
                        "change_pct": round(dev, 1),
                        "detail": f"计划 {m['plan']:,} / 实际 {m['actual']:,}",
                        "direction": "down",
                    })
                elif dev > 50:
                    anomalies.append({
                        "type": "出货超量",
                        "severity": "🟡 预警",
                        "customer": c["customer"],
                        "month": f"{m['month']}月",
                        "change_pct": round(dev, 1),
                        "detail": f"计划 {m['plan']:,} / 实际 {m['actual']:,}",
                        "direction": "up",
                    })

    # H1完成率异常
    for c in qty_data["customers"]:
        if c["h1_plan"] > 10000 and c["completion_pct"] < 60:
            anomalies.append({
                "type": "H1完成率",
                "severity": "🔴 严重",
                "customer": c["customer"],
                "month": "H1",
                "change_pct": round(c["completion_pct"] - 100, 1),
                "detail": f"完成率 {c['completion_pct']}%",
                "direction": "down",
            })

    anomalies.sort(key=lambda a: (0 if "严重" in a["severity"] else 1, -abs(a["change_pct"])))
    return anomalies[:25]


def build_analysis(rev_path: str, qty_path: str) -> dict:
    """完整分析管线"""
    rev = parse_revenue_excel(rev_path)
    qty = parse_quantity_excel(qty_path)
    anomalies = detect_anomalies(rev, qty)

    # 月度收入
    month_names = ["1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月"]
    monthly_rev = []
    for m in range(12):
        total = sum(c["monthly_rev"][m] for c in rev["customers"])
        monthly_rev.append({"month": month_names[m], "revenue": round(total, 0)})

    # Top 客户
    top_customers = sorted(rev["customers"], key=lambda c: c["total_rev"], reverse=True)[:10]
    total_rev = rev["total"]["rev_2025"]
    for c in top_customers:
        c["share"] = round(c["total_rev"] / total_rev * 100, 1) if total_rev else 0

    # 月度出货
    monthly_ship = []
    if qty["totals"]:
        for m in qty["totals"]:
            monthly_ship.append({
                "month": month_names[m["month"]-1],
                "plan": m["plan"],
                "actual": m["actual"],
                "rate": round(m["actual"]/m["plan"]*100, 1) if m["plan"] else 0,
            })

    # 产品组合
    pmix = []
    for code, label in [("FP","功能机"),("SP","智能机"),("PAD","平板")]:
        if code in qty["product_mix"]:
            total = sum(m["actual"] for m in qty["product_mix"][code])
            pmix.append({"type": label, "qty": total})

    # 出货排名
    qty_rank = sorted(qty["customers"], key=lambda c: c["h1_actual"], reverse=True)[:10]

    # 建议
    recs = _generate_recommendations(rev, qty, anomalies)

    # H1 shipment totals
    h1_plan = sum(m["plan"] for m in qty["totals"]) if qty["totals"] else 0
    h1_actual = sum(m["actual"] for m in qty["totals"]) if qty["totals"] else 0

    return {
        "total": rev["total"],
        "monthly_revenue": monthly_rev,
        "top_customers": top_customers,
        "categories": rev["categories"],
        "regions": rev["regions"],
        "anomalies": anomalies,
        "monthly_shipments": monthly_ship,
        "product_mix": pmix,
        "qty_ranking": qty_rank,
        "recommendations": recs,
        "h1_plan": h1_plan,
        "h1_actual": h1_actual,
        "h1_completion": round(h1_actual/h1_plan*100, 1) if h1_plan else 0,
    }


def _generate_recommendations(rev, qty, anomalies):
    recs = []
    hmd = next((c for c in rev["categories"] if c["category"] == "HMD"), None)
    if hmd and hmd["growth_pct"] < 0:
        recs.append(f"⚠️ HMD 收入同比 {hmd['growth_pct']:+.1f}%, 下降 {abs(hmd['growth_amt']):,.0f}万 — 建议确认订单前景及竞品动态")
    for cat in rev["categories"]:
        if cat["growth_pct"] > 50:
            recs.append(f"📈 {cat['category']} 同比增长 +{cat['growth_pct']:.0f}% — 建议确认产能/供应链能否支撑")
    misses = [c for c in qty["customers"] if c["completion_pct"] < 70 and c["h1_plan"] > 50000]
    if misses:
        recs.append(f"📉 {', '.join(c['customer'] for c in misses[:3])} H1出货远低于计划 — 建议重新评估目标")
    if rev["regions"]:
        top_r = rev["regions"][0]
        total_r = sum(r["rev"] for r in rev["regions"])
        if total_r > 0 and top_r["rev"]/total_r > 0.5:
            recs.append(f"🌍 区域集中度过高: {top_r['region']}占比 {top_r['rev']/total_r*100:.0f}% — 建议加大其他区域开拓")
    recs.append("📊 建议启动H2旺季备货计划, 结合历史季节性数据优化排产节奏")
    return recs
